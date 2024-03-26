import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from sklearn import svm
from sklearn.decomposition import PCA
import xlwt
import openpyxl as op
# 1、数据理解和清洗：
#设置数据文件路径：
train_path = "train.csv"
test_path = "test.xlsx"
#读取数据
Data_train = pd.read_csv(train_path)
Data_test = pd.read_excel(test_path)
# 打印属性的数据量和缺失值情况
print(Data_train.info())
print(Data_test.info())

# 打印各属性的基本信息包括计数、均值、标准差、最小值、25%分位数、50%分位数（中位数）、75%分位数和最大值
pd.set_option('display.max_columns',None)
print(Data_train.describe())
print(Data_test.describe())

# 打印分类情况
# 显示中文标题
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False
print(Data_train['fetal_health'].value_counts())

 # 绘图
fig = plt.figure()
# 基线值分布
plt.subplot2grid((3, 3), (0, 0))
#绘制基线值属性的直方图
Data_train['baseline value'].hist()
plt.xlabel(u'基线值 ')
#绘制加速度属性的直方图
plt.subplot2grid((3, 3), (0, 1))
Data_train['accelerations'].hist()
plt.xlabel(u'加速度 ')
#绘制异常短期变量属性的直方图
plt.subplot2grid((3, 3), (0, 2))
Data_train['abnormal_short_term_variability'].hist()
plt.xlabel(u'异常短期变量   ')
#绘制子宫收缩属性的直方图
plt.subplot2grid((3, 3), (2, 0))
Data_train['uterine_contractions'].hist()
plt.xlabel(u'子宫收缩 ')
#绘制短期变异的平均值属性的直方图
plt.subplot2grid((3, 3), (2, 1))
Data_train['mean_value_of_short_term_variability'].hist()
plt.xlabel(u'短期变异的平均值   ')
#绘制零的直方图数量的直方图
plt.subplot2grid((3, 3), (2, 2))
Data_train['histogram_number_of_zeroes'].value_counts().plot(kind='bar')
plt.xlabel(u'零的直方图数量 ')
plt.show()
# 特征提取
# 训练集标签值提取
y_train = np.array(Data_train)
y_train = np.mat(y_train)
y_train1 = y_train[:,22]
y_train1 = np.array(y_train1)
# 训练集特征值提取
x_train = np.array(Data_train)
x_train1 = np.mat(x_train)
x_train1 = x_train[:,1:22]
# 测试集特征值提取
x_test = np.array(Data_test)
x_test1 = np.mat(x_test)
x_test1 = x_test[:,1:22]
# SVM分类器参数设置
#选择线性核函数linear，处理线性可分的问题。选择ovr，用一对其余的方法来处理多类别分类问题。
clf = svm.SVC(C=1,  # 误差项惩罚系数,默认值是1
                  kernel='linear',  # 线性核
                  decision_function_shape='ovr')  # 决策函数

# 模型训练函数接受三个参数：clf表示分类器模型，x_train表示训练集的特征向量，y_train表示训练集的标签。调用分类器的fit方法，将训练集的特征向量和目标值作为参数进行模型训练。
def train(clf, x_train, y_train):
    clf.fit(x_train,  # 训练集特征向量
            y_train.ravel())  # 训练集目标值

# 训练SVM模型
train(clf, x_train1, y_train1)

# 输出训练集的准确率
print('训练集准确率:%.3f' %(clf.score(x_train1, y_train1)))
#预测测试集的目标值，并保存结果到test_predict，再打印结果以及各个样本结果的数量
test_predict = clf.predict(x_test1)
print('具体结果如下')
print(test_predict)
# 将预测结果写入到指定excel中
workbook = xlwt.Workbook(encoding='utf-8')
bg = op.load_workbook(r"predict.xlsx")
sheet = bg["Sheet1"]
for i in range(len(test_predict)):
    sheet.cell(i+2,2,test_predict[i])
bg.save(r"predict.xlsx")